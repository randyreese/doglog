"""
Import historical milestone data from the Excel export into doglog.

Usage:
    python scripts/import_milestones.py <path_to_excel> [--dev]

    --dev  Post to http://localhost:8001/doglog instead of https://mint.local/doglog

The Excel file must have two sheets: "Tess" and "Pickles".
Each sheet has the structure:
    Row 1: Born date in column B
    Row 2: Header row (ignored — values are Excel formulas on a fresh export)
    Row 3+: Wk# | Date | Month | (empty) | Weight | Events | Training

Classification heuristics (Events column → event_type):
    VET keywords  → "vet"
    TRAVEL keywords → "travel"
    Training column → "train"
    Everything else → "life"

Two records are produced when both Events and Training have content on the same row.
A weight-only row (no Events, no Training) produces a "life" record.
"""

import sys
import argparse
from pathlib import Path
import openpyxl
import httpx

# ── keyword sets ──────────────────────────────────────────────────────────────

_VET_KEYWORDS = {
    "vax", "vet", "lac", "vca", "dx", "meds", "rx", "spay", "neuter",
    "glands", "simparica", "lepto", "lyme", "antibiotics", "xray", "cytopoint",
    "ointment", "booster", "boosters", "wellness", "pre-op", "ultrasound",
    "recheck", "traz", "trazodone", "kennel cough", "diarhea", "diarrhea",
    "giardia", "hernia", "flu #", "flu#",
}

_TRAVEL_KEYWORDS = {
    "clintonville", "obx", "prarie oaks", "prairie oaks", "overnight",
    "r/t to", "road trip",
}


def _parse_weight(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        # Strip trailing non-numeric chars (e.g. '48?')
        cleaned = "".join(c for c in str(val) if c.isdigit() or c == ".")
        try:
            return float(cleaned) if cleaned else None
        except ValueError:
            return None


def _classify(events_text: str) -> str:
    lower = events_text.lower()
    for kw in _TRAVEL_KEYWORDS:
        if kw in lower:
            return "travel"
    for kw in _VET_KEYWORDS:
        if kw in lower:
            return "vet"
    return "life"


# ── Excel parsing ─────────────────────────────────────────────────────────────

def _parse_sheet(ws, dog_name: str) -> list[dict]:
    """Return a list of milestone dicts from one dog's sheet."""
    born_date = ws.cell(1, 2).value  # B1
    if not born_date:
        raise ValueError(f"Sheet '{dog_name}': no birthdate in B1")

    records = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        wk, date_val, _month, _empty, weight, events, training = (list(row) + [None] * 7)[:7]

        # Skip rows with no useful data
        if not weight and not events and not training:
            continue
        # Skip the header row that was re-exported as values
        if wk == "Wks" or events == "Events":
            continue
        # date_val may be a datetime or None
        if not date_val:
            continue

        date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val)[:10]

        # Events record
        if events:
            event_type = _classify(str(events))
            records.append({
                "dog_name": dog_name,
                "date": date_str,
                "event_type": event_type,
                "notes1": str(events)[:60],
                "notes2": None,
                "weight_lbs": _parse_weight(weight),
            })
            # Training as second record on same row (notes2 not used — separate record)
            if training:
                records.append({
                    "dog_name": dog_name,
                    "date": date_str,
                    "event_type": "train",
                    "notes1": str(training)[:60],
                    "notes2": None,
                    "weight_lbs": None,
                })
        elif training:
            records.append({
                "dog_name": dog_name,
                "date": date_str,
                "event_type": "train",
                "notes1": str(training)[:60],
                "notes2": None,
                "weight_lbs": None,
            })
        elif weight:
            # Weight-only row → life event
            records.append({
                "dog_name": dog_name,
                "date": date_str,
                "event_type": "life",
                "notes1": None,
                "notes2": None,
                "weight_lbs": float(weight),
            })

    return records


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import milestone data from Excel")
    parser.add_argument("excel", help="Path to ConvertMilestonesData.xlsx")
    parser.add_argument("--dev", action="store_true", help="Use localhost:8001 backend")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, do not POST")
    args = parser.parse_args()

    base_url = "http://localhost:8001/doglog" if args.dev else "https://mint.local/doglog"

    print(f"Loading {args.excel} ...")
    wb = openpyxl.load_workbook(args.excel, data_only=True)

    all_records: list[dict] = []
    for sheet_name in ["Tess", "Pickles"]:
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not found — skipping")
            continue
        recs = _parse_sheet(wb[sheet_name], sheet_name)
        print(f"  {sheet_name}: {len(recs)} records parsed")
        all_records.extend(recs)

    print(f"\nTotal: {len(all_records)} records")

    if args.dry_run:
        for r in all_records:
            print(r)
        return

    # Fetch dog list to resolve names → IDs
    try:
        import truststore
        truststore.inject_into_ssl()
    except ImportError:
        pass

    client = httpx.Client(base_url=base_url, timeout=15)

    try:
        dogs = client.get("/dogs/").raise_for_status().json()
    except Exception as e:
        print(f"ERROR: could not fetch dogs list: {e}")
        sys.exit(1)

    dog_id_map = {d["name"]: d["id"] for d in dogs}
    print(f"Dogs found: {dog_id_map}")

    ok = 0
    errors = 0
    for r in all_records:
        dog_name = r.pop("dog_name")
        r["dog_id"] = dog_id_map.get(dog_name)
        if r["dog_id"] is None:
            print(f"  WARNING: dog '{dog_name}' not found in DB — setting to All (NULL)")

        try:
            client.post("/milestones/", json=r).raise_for_status()
            ok += 1
        except Exception as e:
            print(f"  ERROR posting {r}: {e}")
            errors += 1

    print(f"\nDone: {ok} imported, {errors} errors")


if __name__ == "__main__":
    main()
