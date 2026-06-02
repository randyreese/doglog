"""
Import Sprint 9 historical meal/health/medication data from Sprint9input.xlsx.

Usage:
    python scripts/import_history.py --file <path> [options]

Options:
    --file      Path to Sprint9input.xlsx (required)
    --tabs      Comma-separated tab names (default: Jan,Feb,Mar,Apr,May)
    --dog       "Tess", "Pickles", or "all" (default: all)
    --dry-run   Parse and count without writing anything
    --force     Skip purge check (script will skip duplicates)
    --purge     Delete all import-range records before running (handles dummy test data)
    --db        Path to SQLite DB file (default: backend/doglog.db)
"""

import os
import sys
import argparse
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl

IMPORT_START = date(2026, 1, 1)
IMPORT_END   = date(2026, 5, 31)

# Excel column indices (1-based, openpyxl)
COL_DATE      = 2   # B
COL_DOG       = 3   # C
COL_BFAST     = 5   # E
COL_LUNCH     = 6   # F
COL_DINNER    = 7   # G
COL_PM_SNACK  = 8   # H
COL_VOMIT     = 12  # L
COL_NOTES     = 13  # M

SLOT_COLS = [
    ("breakfast", COL_BFAST),
    ("lunch",     COL_LUNCH),
    ("dinner",    COL_DINNER),
    ("pm_snack",  COL_PM_SNACK),
]


# ── DB helpers ────────────────────────────────────────────────────────────────

def load_meal_configs(db, models):
    """Return {(dog_id, slot): [(effective_date, [food_names]), ...]} sorted asc by date."""
    result = {}
    for cfg in db.query(models.MealConfig).all():
        items = (
            db.query(models.MealConfigItem)
            .filter_by(meal_config_id=cfg.id)
            .order_by(models.MealConfigItem.sort_order)
            .all()
        )
        key = (cfg.dog_id, cfg.meal_slot)
        result.setdefault(key, []).append((cfg.effective_date, [i.food_name for i in items]))
    for key in result:
        result[key].sort(key=lambda x: x[0])
    return result


def find_config(configs, dog_id, slot, target_date):
    """Return [food_names] for the config active on target_date, or None."""
    active = None
    for eff_date, foods in configs.get((dog_id, slot), []):
        if eff_date <= target_date:
            active = foods
        else:
            break
    return active


def load_sucralfate(db, models):
    """Return (medication_id, am_dose_label, pm_dose_label) or raise RuntimeError."""
    med = db.query(models.Medication).filter(models.Medication.name.ilike("%Sucralfate%")).first()
    if not med:
        raise RuntimeError("Sucralfate medication not found in DB")
    doses = (
        db.query(models.MedicationDose)
        .filter_by(medication_id=med.id)
        .order_by(models.MedicationDose.sort_order)
        .all()
    )
    if len(doses) < 2:
        raise RuntimeError(f"Sucralfate: expected ≥2 doses, found {len(doses)}")
    return med.id, doses[0].label, doses[1].label


# ── upsert functions ──────────────────────────────────────────────────────────

def upsert_meal_log(db, models, dog_id, slot, meal_date, pct, foods, dry_run, counts):
    existing = db.query(models.MealLog).filter_by(dog_id=dog_id, slot=slot, meal_date=meal_date).first()
    if existing:
        counts["meal_skipped"] += 1
        return
    if not dry_run:
        db.add(models.MealLog(
            dog_id=dog_id,
            slot=slot,
            meal_date=meal_date,
            percent_consumed=pct,
            notes=None,
            ingredients=json.dumps({f: True for f in foods}),
        ))
    counts["meal_created"] += 1


def upsert_vomit(db, models, dog_id, event_date, dry_run, counts):
    day_start = datetime(event_date.year, event_date.month, event_date.day, 0, 0, 0)
    day_end   = datetime(event_date.year, event_date.month, event_date.day, 23, 59, 59)
    existing = (
        db.query(models.OtherEvent)
        .filter(
            models.OtherEvent.dog_id == dog_id,
            models.OtherEvent.type == "vomit",
            models.OtherEvent.timestamp >= day_start,
            models.OtherEvent.timestamp <= day_end,
        )
        .first()
    )
    if existing:
        counts["vomit_skipped"] += 1
        return
    if not dry_run:
        db.add(models.OtherEvent(
            dog_id=dog_id,
            timestamp=datetime(event_date.year, event_date.month, event_date.day, 12, 0, 0),
            type="vomit",
        ))
    counts["vomit_created"] += 1


def upsert_medication_log(db, models, dog_id, med_id, log_date, dose_labels, dry_run, counts):
    existing = (
        db.query(models.MedicationLog)
        .filter_by(dog_id=dog_id, medication_id=med_id, log_date=log_date)
        .first()
    )
    if existing:
        counts["med_skipped"] += 1
        return
    if not dry_run:
        db.add(models.MedicationLog(
            dog_id=dog_id,
            medication_id=med_id,
            log_date=log_date,
            doses_given=json.dumps(dose_labels),
        ))
    counts["med_created"] += 1


# ── parsing helpers ───────────────────────────────────────────────────────────

def parse_pct(val):
    if val is None:
        return None
    try:
        return int(round(float(val)))
    except (ValueError, TypeError):
        return None


def parse_sucralfate(notes_val, am_label, pm_label):
    """Return list of dose labels, or None if not applicable."""
    if not notes_val:
        return None
    s = str(notes_val).lower()
    if "no sucralfate" in s:
        return None
    if "sucralfate" not in s:
        return None
    doses = []
    if "am" in s:
        doses.append(am_label)
    if "pm" in s:
        doses.append(pm_label)
    return doses or None


def to_date(val):
    if val is None:
        return None
    if hasattr(val, "date"):
        return val.date()
    if isinstance(val, date):
        return val
    return None


# ── tab processor ─────────────────────────────────────────────────────────────

def process_tab(ws, tab_name, dog_filter, dog_id_map, configs,
                succ_med_id, am_label, pm_label,
                db, models, dry_run, counts):

    col_m_header = ws.cell(4, COL_NOTES).value
    is_notes_tab = col_m_header and "notes" in str(col_m_header).lower()
    print(f"  [{tab_name}] col M = '{col_m_header}'  notes_tab={is_notes_tab}")

    for row_idx in range(5, ws.max_row + 1):
        date_val = ws.cell(row_idx, COL_DATE).value
        dog_cell = ws.cell(row_idx, COL_DOG).value

        if not date_val or not dog_cell:
            continue

        dog_name = str(dog_cell).strip()
        if dog_filter != "all" and dog_name.lower() != dog_filter.lower():
            continue

        dog_id = dog_id_map.get(dog_name)
        if dog_id is None:
            print(f"    WARNING row {row_idx}: unknown dog '{dog_name}' — skipping")
            continue

        row_date = to_date(date_val)
        if row_date is None or not (IMPORT_START <= row_date <= IMPORT_END):
            continue

        # Explicit meal slots
        for slot, col in SLOT_COLS:
            pct = parse_pct(ws.cell(row_idx, col).value)
            if pct is None:
                counts["meal_no_value"] += 1
                continue
            foods = find_config(configs, dog_id, slot, row_date)
            if foods is None:
                counts["meal_no_config"] += 1
                continue
            upsert_meal_log(db, models, dog_id, slot, row_date, pct, foods, dry_run, counts)

        # AM snack — synthesized at 100%
        am_foods = find_config(configs, dog_id, "am_snack", row_date)
        if am_foods is not None:
            upsert_meal_log(db, models, dog_id, "am_snack", row_date, 100, am_foods, dry_run, counts)

        # Vomit (col L)
        vomit_val = ws.cell(row_idx, COL_VOMIT).value
        if vomit_val and str(vomit_val).strip().lower() == "x":
            upsert_vomit(db, models, dog_id, row_date, dry_run, counts)

        # Sucralfate (col M, Apr–May only)
        if is_notes_tab and succ_med_id:
            doses = parse_sucralfate(ws.cell(row_idx, COL_NOTES).value, am_label, pm_label)
            if doses:
                upsert_medication_log(db, models, dog_id, succ_med_id, row_date, doses, dry_run, counts)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import Sprint 9 historical data")
    parser.add_argument("--file", required=True, help="Path to Sprint9input.xlsx")
    parser.add_argument("--tabs", default="Jan,Feb,Mar,Apr,May")
    parser.add_argument("--dog", default="all", help="'Tess', 'Pickles', or 'all'")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--force", action="store_true", help="Skip purge check")
    parser.add_argument("--purge", action="store_true", help="Delete import-range records before running")
    parser.add_argument("--db", default="backend/doglog.db")
    args = parser.parse_args()

    db_path = Path(args.db).resolve()
    if not db_path.exists():
        print(f"ERROR: DB not found at {db_path}")
        sys.exit(1)

    excel_path = Path(args.file)
    if not excel_path.exists():
        print(f"ERROR: Excel file not found at {excel_path}")
        sys.exit(1)

    # Import backend modules after resolving DB path
    os.environ["DATABASE_URL"] = f"sqlite:///{db_path}"
    sys.path.insert(0, str(Path(__file__).parent.parent / "backend"))
    import models as _models
    from database import SessionLocal

    db = SessionLocal()
    try:
        dogs_list = db.query(_models.Dog).filter_by(active=True).all()
        dog_id_map = {d.name: d.id for d in dogs_list}
        print(f"Dogs: {dog_id_map}")

        # Purge
        if args.purge and not args.dry_run:
            n_meals = db.query(_models.MealLog).filter(
                _models.MealLog.meal_date >= IMPORT_START,
                _models.MealLog.meal_date <= IMPORT_END,
            ).delete()
            n_vomit = db.query(_models.OtherEvent).filter(
                _models.OtherEvent.type == "vomit",
                _models.OtherEvent.timestamp >= datetime(2026, 1, 1),
                _models.OtherEvent.timestamp < datetime(2026, 6, 1),
            ).delete()
            n_meds = db.query(_models.MedicationLog).filter(
                _models.MedicationLog.log_date >= IMPORT_START,
                _models.MedicationLog.log_date <= IMPORT_END,
            ).delete()
            db.commit()
            print(f"Purged: {n_meals} meal_logs, {n_vomit} vomit events, {n_meds} medication_logs")

        # Purge check
        if not args.force and not args.purge:
            n = db.query(_models.MealLog).filter(
                _models.MealLog.meal_date >= IMPORT_START,
                _models.MealLog.meal_date <= IMPORT_END,
            ).count()
            if n > 0:
                print(f"ERROR: {n} meal_log records already exist in 1/1–5/31 range.")
                print("Use --purge to delete them first, or --force to skip duplicates.")
                sys.exit(1)

        configs = load_meal_configs(db, _models)
        print(f"Meal configs: {len(configs)} (dog, slot) pairs")

        try:
            succ_med_id, am_label, pm_label = load_sucralfate(db, _models)
            print(f"Sucralfate: med_id={succ_med_id}, am='{am_label}', pm='{pm_label}'")
        except RuntimeError as e:
            print(f"WARNING: {e} — medication logs will be skipped")
            succ_med_id = am_label = pm_label = None

        counts = {
            "meal_created": 0, "meal_skipped": 0,
            "meal_no_value": 0, "meal_no_config": 0,
            "vomit_created": 0, "vomit_skipped": 0,
            "med_created": 0,   "med_skipped": 0,
        }

        wb = openpyxl.load_workbook(args.file, data_only=True)
        tabs = [t.strip() for t in args.tabs.split(",")]

        for tab in tabs:
            if tab not in wb.sheetnames:
                print(f"WARNING: tab '{tab}' not found in workbook — skipping")
                continue
            process_tab(
                wb[tab], tab, args.dog, dog_id_map, configs,
                succ_med_id, am_label, pm_label,
                db, _models, args.dry_run, counts,
            )

        if not args.dry_run:
            db.commit()
            print("\nCommitted to DB.")
        else:
            db.rollback()
            print("\n[DRY RUN — nothing written]")

        print(f"""
Summary
-------
  Meal logs created:         {counts['meal_created']}
  Meal logs skipped (dupe):  {counts['meal_skipped']}
  Slots skipped (no value):  {counts['meal_no_value']}
  Slots skipped (no config): {counts['meal_no_config']}
  Vomit events created:      {counts['vomit_created']}
  Vomit events skipped:      {counts['vomit_skipped']}
  Medication logs created:   {counts['med_created']}
  Medication logs skipped:   {counts['med_skipped']}
""")

    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
