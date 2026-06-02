"""Vet report generator — writes one month of data to a named-range data zone in an xlsx file."""
import calendar
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import httpx
import openpyxl


def _fmt_meal(pct, notes: str | None) -> str:
    if pct is None:
        return ""
    notes = (notes or "").strip()
    if pct == 0:
        return f"*X* {notes}".strip() if notes else "X"
    display = "ok" if pct == 100 else str(pct)
    return f"*{display}* {notes}".strip() if notes else display


def _fmt_health(events: list) -> str:
    parts = []
    for ev in sorted(events, key=lambda e: e["timestamp"]):
        ts = datetime.fromisoformat(ev["timestamp"])
        h = ts.strftime("%I").lstrip("0") or "12"
        time_str = f"{h}:{ts.strftime('%M%p').lower()}"
        label = ev["_label"]
        notes = (ev.get("notes") or "").strip()
        parts.append(f"{time_str} {label}" + (f" – {notes}" if notes else ""))
    return "\n".join(parts)


def _fmt_medications(logs_for_date: list, active_meds: list) -> str:
    if not active_meds:
        return ""
    log_by_med = {log["medication_id"]: log for log in logs_for_date}
    lines = []
    for med in active_meds:
        total = len(med["doses"])
        log = log_by_med.get(med["id"])
        if log is None:
            continue
        given = log["doses_given"]
        if not given:
            status = "Skipped"
        elif len(given) >= total:
            status = "All given"
        else:
            status = f"{len(given)} of {total} given"
        if len(active_meds) > 1:
            lines.append(f"{med['name']}: {status}")
        else:
            lines.append(status)
    return "\n".join(lines)


def _find_anchor(wb: openpyxl.Workbook, ws) -> tuple[int, int]:
    """Return (row, col) of the data_anchor for this sheet.

    Looks for data_anchor_{sheet_lower} (e.g. data_anchor_apr) then
    falls back to the generic data_anchor. All names must be workbook-scoped.
    """
    key = f"data_anchor_{ws.title.lower()}"
    dn = wb.defined_names.get(key) or wb.defined_names.get("data_anchor")
    if dn is None:
        raise ValueError(
            f"Named range '{key}' (or 'data_anchor') not found. "
            f"Available names: {', '.join(wb.defined_names.keys()) or '(none)'}"
        )
    _, ref = list(dn.destinations)[0]
    cell = ws[ref]
    return cell.row, cell.column


def generate(
    dog_id: int,
    month: int,
    year: int,
    output_file: str,
    base_url: str = "http://localhost:8001",
) -> None:
    client = httpx.Client(base_url=base_url, timeout=30)

    def _get(path, **params):
        r = client.get(path, params=params if params else None)
        r.raise_for_status()
        return r.json()

    # Reference data
    dogs = {d["id"]: d for d in _get("/dogs/")}
    if dog_id not in dogs:
        raise ValueError(f"Dog id={dog_id} not found")

    health_types = {ht["value"]: ht for ht in _get("/health-types")}
    slots = [s["value"] for s in _get("/meal-slots")]

    # Date range
    _, days_in_month = calendar.monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, days_in_month)

    # Meal logs
    meal_raw = _get("/meal-logs/range/", dog_id=dog_id,
                    start_date=str(start), end_date=str(end))
    meal_by_date_slot = {(log["meal_date"], log["slot"]): log for log in meal_raw}

    # Health events
    health_raw = _get("/health-events/", dog_id=dog_id,
                      since=f"{start}T00:00:00", until=f"{end}T23:59:59", limit=1000)
    for ev in health_raw:
        ht = health_types.get(ev["type"], {})
        ev["_label"] = ht.get("label", ev["type"])
        ev["_report_column"] = ht.get("report_column", "")
    health_by_date: dict[str, list] = defaultdict(list)
    for ev in health_raw:
        ev_date = str(datetime.fromisoformat(ev["timestamp"]).date())
        health_by_date[ev_date].append(ev)

    # Medications active during this month
    meds_raw = _get("/medications/", dog_id=dog_id)
    active_meds = []
    for med in meds_raw:
        med_start = date.fromisoformat(med["start_date"]) if med.get("start_date") else date.min
        med_end = date.fromisoformat(med["end_date"]) if med.get("end_date") else date.max
        if med_start <= end and med_end >= start:
            active_meds.append(med)

    # Medication logs for the month
    med_logs_raw = _get("/medication-logs/range/", dog_id=dog_id,
                        start_date=str(start), end_date=str(end))
    med_logs_by_date: dict[str, list] = defaultdict(list)
    for log in med_logs_raw:
        med_logs_by_date[log["log_date"]].append(log)

    # Build one row per day
    rows = []
    for day_num in range(1, days_in_month + 1):
        d = date(year, month, day_num)
        d_str = str(d)

        meal_cells = []
        for slot in slots:
            log = meal_by_date_slot.get((d_str, slot))
            meal_cells.append(_fmt_meal(log["percent_consumed"], log.get("notes")) if log else "")

        day_events = health_by_date.get(d_str, [])
        activity_str = _fmt_health([e for e in day_events if e["_report_column"] == "activity"])
        event_str = _fmt_health([e for e in day_events if e["_report_column"] == "event"])

        meds_today = [
            m for m in active_meds
            if (date.fromisoformat(m["start_date"]) if m.get("start_date") else date.min) <= d
            and (date.fromisoformat(m["end_date"]) if m.get("end_date") else date.max) >= d
        ]
        med_str = _fmt_medications(med_logs_by_date.get(d_str, []), meds_today)

        rows.append([d] + meal_cells + [activity_str, event_str, med_str])

    # Open workbook
    output_file = str(output_file)
    wb = openpyxl.load_workbook(output_file)
    sheet_name = start.strftime("%b")  # "Jan", "Feb", …
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {Path(output_file).name}. "
            f"Available sheets: {', '.join(wb.sheetnames)}"
        )
    ws = wb[sheet_name]

    anchor_row, anchor_col = _find_anchor(wb, ws)

    # Row 0: dog name; Row 1: period (presentation grid rows 1–2)
    ws.cell(row=anchor_row, column=anchor_col, value=dogs[dog_id]["name"])
    ws.cell(row=anchor_row + 1, column=anchor_col, value=start.strftime("%b %Y"))

    # Rows 4+: day data (presentation grid row 5+)
    for r_idx, row_data in enumerate(rows):
        for c_idx, value in enumerate(row_data):
            ws.cell(row=anchor_row + 4 + r_idx, column=anchor_col + c_idx, value=value)

    wb.save(output_file)
